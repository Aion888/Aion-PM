from pathlib import Path
import re
from openpyxl import load_workbook

XLSX = Path("data/Project_Aion_PM_System.xlsx")
SPEC_DIR = Path("Project_Aion/01_Project_Framework/00_Master_Index/Framework_Directory_Spec")

TICKETS_SHEET = "04_Tickets"
VALID_SHEET = "99_Validation"

def guess_header_row(ws, scan_rows=30, max_cols=80) -> int:
    best_row = 1
    best = (-1, -1)
    for r in range(1, min(scan_rows, ws.max_row) + 1):
        vals = []
        for c in range(1, max_cols + 1):
            v = ws.cell(r, c).value
            vals.append("" if v is None else str(v).strip())
        while vals and vals[-1] == "":
            vals.pop()
        nonempty = sum(v != "" for v in vals)
        if nonempty == 0:
            continue
        stringy = sum(v != "" and not re.fullmatch(r"[-+]?\d+(\.\d+)?", v) for v in vals)
        score = (nonempty, stringy)
        if score > best:
            best = score
            best_row = r
    return best_row

def col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def main():
    wb = load_workbook(XLSX)
    ws = wb[TICKETS_SHEET]

    # Ensure validation sheet exists
    vws = wb[VALID_SHEET] if VALID_SHEET in wb.sheetnames else wb.create_sheet(VALID_SHEET)

    # Read realms from 99_Validation col A, or fallback to spec filenames
    realms_in_sheet = []
    r = 2
    while True:
        v = vws[f"A{r}"].value
        if v is None or str(v).strip() == "":
            break
        realms_in_sheet.append(str(v).strip())
        r += 1

    if not realms_in_sheet:
        realms_in_sheet = sorted([p.stem for p in SPEC_DIR.glob("*.md")])
        vws["A1"] = "Allowed_Realms"
        for i, realm in enumerate(realms_in_sheet, start=2):
            vws[f"A{i}"] = realm

    # Build realm -> framework path mapping in col C
    vws["C1"] = "Framework_Path_By_Realm"
    for i, realm in enumerate(realms_in_sheet, start=2):
        # Special-case: triage lives at 00_Triage_Inbox folder, not Project_Aion/00_Triage_Inbox.md
        if realm == "00_Triage_Inbox":
            vws[f"C{i}"] = "Project_Aion/00_Triage_Inbox"
        else:
            vws[f"C{i}"] = f"Project_Aion/{realm}"

    last = len(realms_in_sheet) + 1

    # Locate columns in tickets
    hdr = guess_header_row(ws)
    col = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hdr, c).value
        if v:
            col[str(v).strip()] = c

    needed = ["Realm", "Framework_Path", "Roadmap_Milestone", "Start"]
    missing = [n for n in needed if n not in col]
    if missing:
        raise SystemExit(f"Missing columns in {TICKETS_SHEET} header row {hdr}: {missing}")

    realm_col = col_letter(col["Realm"])
    path_col  = col_letter(col["Framework_Path"])
    mile_col  = col_letter(col["Roadmap_Milestone"])
    start_col = col_letter(col["Start"])

    for row in range(hdr + 1, ws.max_row + 1):
        realm_cell = f"{realm_col}{row}"
        start_cell = f"{start_col}{row}"

        # Only fill blanks
        if ws[f"{path_col}{row}"].value in (None, ""):
            ws[f"{path_col}{row}"].value = (
                f'=IF({realm_cell}="","",IFERROR('
                f'XLOOKUP({realm_cell},\'{VALID_SHEET}\'!$A$2:$A${last},\'{VALID_SHEET}\'!$C$2:$C${last},""),""))'
            )

        if ws[f"{mile_col}{row}"].value in (None, ""):
            ws[f"{mile_col}{row}"].value = (
                f'=IF({start_cell}="","", "Q"&ROUNDUP(MONTH({start_cell})/3,0)&"_"&TEXT({start_cell},"yyyy"))'
            )

    wb.save(XLSX)
    print("✅ Applied autofill formulas for Framework_Path + Roadmap_Milestone (blank cells only).")

if __name__ == "__main__":
    main()
